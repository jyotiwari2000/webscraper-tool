from flask import Flask, render_template, request
import requests
import openpyxl
from lxml import html
from openpyxl.styles import NamedStyle, Font

app = Flask(__name__)

def setup_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    style = NamedStyle(name="bold_font", font=Font(bold=True))
    for col_num, header in enumerate(["URLs", "Label","Status", "Server", "Redirected URLs", "Status", "Server"], start=1):
        sheet.cell(row=1, column=col_num, value=header).style = style

    return workbook, sheet

def get_https_links(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        tree = html.fromstring(response.content)
        
        # Filter links by button class "btn-title"
        https_links = tree.xpath('//a[starts-with(@href, "https://")]/@href')

        return https_links

    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

def get_https_links_text(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        tree = html.fromstring(response.content)
        
        # Extract links and corresponding span text
        links = tree.xpath('//a[starts-with(@href, "https://")]')
        
        url_label_pairs = []
        for link in links:
            href = link.get('href')
            span_text = link.xpath('.//span/text()')
            if span_text:
                label = span_text[0]
            else:
                label = 'No Label Available'
            url_label_pairs.append((href, label))
        
        return url_label_pairs

    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scrape', methods=['POST'])
def scrape():
    url_to_scrape = request.form['url']
    https_links = get_https_links(url_to_scrape)
    https_links_text = get_https_links_text(url_to_scrape)
    
    results = []
    
    if https_links:
        workbook, sheet = setup_workbook()
        row_number = 2

        for url in https_links:
            sheet.cell(row=row_number, column=1, value=url)
            row_number += 1

        if https_links_text:
            row_number = 2
            for url, label in https_links_text:
                sheet.cell(row=row_number, column=1, value=url)
                sheet.cell(row=row_number, column=2, value=label)
                row_number += 1

        workbook.save("webdata_updated_New.xlsx")
        results.append("URLs and labels saved to webdata_updated_New.xlsx")

        workbook = openpyxl.load_workbook("webdata_updated_New.xlsx")
        sheet = workbook.active

        for row_number in range(2, sheet.max_row + 1):
            url_cell = sheet.cell(row=row_number, column=1)
            url = url_cell.value

            try:
                with requests.Session() as session:
                    response = session.get(url, allow_redirects=True)

                status_code = f"{response.history[0].status_code if response.history else response.status_code}"
                server_code = f"{response.history[0].headers.get('Server') if response.history else response.headers.get('Server')}"

                sheet.cell(row=row_number, column=3, value=str(status_code))
                sheet.cell(row=row_number, column=4, value=server_code)
                sheet.cell(row=row_number, column=5, value=response.url)

                if response.history:
                    for redirected_response in response.history:
                        redirected_status = f"{redirected_response.status_code}"
                        redirected_server = f"{redirected_response.headers.get('Server')}"
                        sheet.cell(row=row_number, column=6, value=str(redirected_status))
                        sheet.cell(row=row_number, column=7, value=redirected_server)
                        row_number += 1

            except requests.exceptions.ConnectionError as e:
                if hasattr(e, 'response') and e.response is not None:
                    results.append(f"Connection error for URL: {url}. Status code: {e.response.status_code}")
                else:
                    results.append(f"Connection error for URL: {url}. Error: {e}")

        workbook.save("webdata_updated_New.xlsx")
        results.append("Updated data saved to webdata_updated_New.xlsx")

    else:
        results.append("No HTTPS links found on the webpage.")

    # Read the contents of the Excel file
    workbook = openpyxl.load_workbook("webdata_updated_New.xlsx")
    sheet = workbook.active
    excel_data = []
    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        excel_data.append(row)

    return render_template('results.html', results=results, excel_data=excel_data)

if __name__ == "__main__":
    app.run(debug=True)

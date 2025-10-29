import requests
import openpyxl
from lxml import html
from openpyxl.styles import NamedStyle, Font

def setup_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    style = NamedStyle(name="bold_font", font=Font(bold=True))
    for col_num, header in enumerate(["URLs", "Label","Status", "Server", "Redirected URLs", "Status", "Server"], start=1):
        sheet.cell(row=1, column=col_num, value=header).style = style

    return workbook, sheet

def get_https_links(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        tree = html.fromstring(response.content)
        
        # Filter links by button class "btn-title"
        https_links = tree.xpath('//a[starts-with(@href, "https://")]/@href')

        return https_links

    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None
def get_https_links_text(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        tree = html.fromstring(response.content)
        
        # Extract links and corresponding span text
        links = tree.xpath('//a[starts-with(@href, "https://")]')
        
        url_label_pairs = []
        for link in links:
            href = link.get('href')
            span_text = link.xpath('.//span/text()')
            if span_text:
                label = span_text[0]
            else:
                label = 'No Label Available'
            url_label_pairs.append((href, label))
        
        return url_label_pairs

    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None





def main():
    try:
        url_to_scrape = input("Enter the URL to scrape: ")

        if url_to_scrape.startswith(('http://', 'https://')):
            https_links = get_https_links(url_to_scrape)
            https_links_text = get_https_links_text(url_to_scrape)

            if https_links:
                workbook, sheet = setup_workbook()
                row_number = 2

                for url in https_links:
                    sheet.cell(row=row_number, column=1, value=url)
                    row_number += 1

                if https_links_text:
                    row_number = 2
                    for url, label in https_links_text:
                        sheet.cell(row=row_number, column=1, value=url)
                        sheet.cell(row=row_number, column=2, value=label)
                        row_number += 1

                workbook.save("webdata_updated_New.xlsx")
                print("URLs and labels saved to webdata_updated_New.xlsx")

                # Load the saved workbook to update status, server, and redirected URLs
                workbook = openpyxl.load_workbook("webdata_updated_New.xlsx")
                sheet = workbook.active

                for row_number in range(2, sheet.max_row + 1):
                    url_cell = sheet.cell(row=row_number, column=1)
                    url = url_cell.value

                    try:
                        with requests.Session() as session:
                            response = session.get(url, allow_redirects=True)

                        status_code = f"{response.history[0].status_code if response.history else response.status_code}"
                        server_code = f"{response.history[0].headers.get('Server') if response.history else response.headers.get('Server')}"

                        # Update status, server, and redirected URLs
                        sheet.cell(row=row_number, column=3, value=str(status_code))
                        sheet.cell(row=row_number, column=4, value=server_code)
                        sheet.cell(row=row_number, column=5, value=response.url)

                        # For redirected URLs
                        if response.history:
                            for redirected_response in response.history:
                                redirected_status = f"{redirected_response.status_code}"
                                redirected_server = f"{redirected_response.headers.get('Server')}"
                                sheet.cell(row=row_number, column=6, value=str(redirected_status))
                                sheet.cell(row=row_number, column=7, value=redirected_server)
                                row_number += 1

                    except requests.exceptions.ConnectionError as e:
                        if hasattr(e, 'response') and e.response is not None:
                            print(f"Connection error for URL: {url}. Status code: {e.response.status_code}")
                        else:
                            print(f"Connection error for URL: {url}. Error: {e}")

                workbook.save("webdata_updated_New.xlsx")
                print("Updated data saved to webdata_updated_New.xlsx")

            else:
                print("No HTTPS links found on the webpage.")
        else:
            print("Invalid URL. Make sure it starts with 'http://' or 'https://'.")
    
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()
